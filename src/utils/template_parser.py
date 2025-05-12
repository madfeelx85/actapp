import re
from openpyxl import load_workbook
from pathlib import Path

FIELD_PATTERN = re.compile(r"{{\s*(\w+)\s*}}")
SUFFIX_PATTERN = re.compile(r"(.*)_(line|part|row|col|chunk)_\d+$")

def normalize_field_name(field: str) -> str:
    match = SUFFIX_PATTERN.match(field)
    return match.group(1) if match else field

def extract_fields_from_excel(file_path: Path) -> list[str]:
    workbook = load_workbook(filename=file_path)
    raw_fields = set()

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    raw_fields.update(FIELD_PATTERN.findall(cell))

    # нормализуем имена
    normalized = {normalize_field_name(f) for f in raw_fields}
    return sorted(normalized)
