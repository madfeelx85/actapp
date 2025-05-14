from pathlib import Path
from io import BytesIO
from datetime import datetime
from loguru import logger
import win32com.client as win32

from openpyxl import load_workbook

from core.models.act import Act
from core.models.build_object import BuildObject
from utils.data_expander import expand_data_for_template


class ActDocumentGenerator:
    def __init__(self, act: Act, build_object: BuildObject):
        self.act = act
        self.build_object = build_object
        self.template_path = Path(self.act.template.path)

    def _fill_template(self):
        workbook = load_workbook(self.template_path)
        sheet = workbook.active

        expanded_data = expand_data_for_template(self.act.data)

        for row in sheet.iter_rows():
            for cell in row:
                if (
                    isinstance(cell.value, str)
                    and "{{" in cell.value
                    and "}}" in cell.value
                ):
                    for key, value in expanded_data.items():
                        placeholder = f"{{{{ {key} }}}}"
                        if placeholder in cell.value:
                            cell.value = cell.value.replace(placeholder, str(value))

        return workbook

    def save_to_file(self) -> Path:
        workbook = self._fill_template()

        output_dir = Path("output") / self.build_object.name
        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y-%m-%d %H-%M-%S")
        act_number = self.act.data.get("act_number", "Не указано")
        filename = f"{self.act.name} №{act_number} от {timestamp}.xlsx"
        output_path = output_dir / filename

        workbook.save(output_path)
        pdf_path = self.convert_to_pdf(output_path)
        logger.info(f"[GENERATE] PDF saved to {pdf_path}")

        return output_path

    def get_as_bytes(self) -> BytesIO:
        workbook = self._fill_template()
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    def convert_to_pdf(self, excel_path: Path) -> Path:
        pdf_path = excel_path.with_suffix(".pdf")

        excel = win32.gencache.EnsureDispatch("Excel.Application")
        excel.Visible = False

        wb = excel.Workbooks.Open(str(excel_path.resolve()))
        try:
            wb.ExportAsFixedFormat(0, str(pdf_path.resolve()))
        finally:
            wb.Close(False)
            excel.Quit()

        return pdf_path
